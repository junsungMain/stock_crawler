import pandas as pd
import concurrent.futures
from tqdm import tqdm
from openpyxl import load_workbook
from module.financial_data import get_financial_extra_data, get_financial_data
from module.main_data import get_stock_extra_data
from module.new_and_disclosure import get_latest_disclosure, get_latest_news
import os
from module.common import retry_on_failure, setup_logging_and_cleanup
import requests
import functools
import logging
from module.stock_list import get_theme_list
from datetime import datetime
import shutil

def process_stock_list(template_excel_path):
    # 템플릿 파일 복사  
    today_str = datetime.now().strftime('%Y%m%d')
    result_excel_path = f"{today_str}_주식현황.xlsx"
    shutil.copy(template_excel_path, result_excel_path)

    # session연결 설정
    session = requests.Session()
    adapter = requests.adapters.HTTPAdapter(
        pool_connections=10,
        pool_maxsize=20,
        max_retries=3,
        pool_block=False
    )
    session.mount('https://', adapter)
    session.mount('http://', adapter)
    session.timeout = (5, 10)
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'ko-KR,ko;q=0.9,en;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive'
    })

    # 종목 코드 리스트 가져오기
    data = get_theme_list(session=session)
    stock_codes = list(data.keys())

    # 코드별 주식 데이터 가져오기
    results = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(os.cpu_count() * 2, 20)) as executor:
        fetcher = functools.partial(fetch_all_data_for_stock, session=session)
        results = list(tqdm(executor.map(fetcher, stock_codes), total=len(stock_codes), desc="전체 데이터 수집 중"))
    session.close()
    
    if not results:
        logging.warning("수집된 데이터가 없습니다.")
        return
    
    for row in results:
        code = row.get('종목코드')
        if code and code in data:
            data[code].update(row)
        elif code:
            data[code] = row

    df_merged = pd.DataFrame(data.values())
    
    wb = load_workbook(result_excel_path)
    ws = wb.active
    headers = {cell.value: cell.column for cell in ws[1]}

    all_data_cols = ['종목코드'] + [col for col in df_merged.columns if col != '종목코드']

    for col_name in all_data_cols:
        if col_name not in headers:
            new_col_idx = ws.max_column + 1
            ws.cell(row=1, column=new_col_idx, value=col_name)
            headers[col_name] = new_col_idx

    for r_idx, row in df_merged.iterrows():
        excel_row_num = r_idx + 2 
        for col_name in all_data_cols:
            if col_name in row:
                target_col_idx = headers.get(col_name)
                if target_col_idx:
                    cell = ws.cell(row=excel_row_num, column=target_col_idx)
                    cell.value = row[col_name]
                    
    wb.save(result_excel_path)
    logging.info(f"데이터가 {result_excel_path}에 저장되었습니다.")

def fetch_all_data_for_stock(stock_code, session):
    
    getters = {
        'stock_extra_data': functools.partial(get_stock_extra_data, session=session),
        'financial_data': functools.partial(get_financial_data, session=session),
        'financial_extra_data': functools.partial(get_financial_extra_data, session=session),
        'latest_news': functools.partial(get_latest_news, session=session),
        'latest_disclosure': functools.partial(get_latest_disclosure, session=session),
    }

    combined_data = {'종목코드': stock_code}
    
    for key, getter in getters.items():
        data = retry_on_failure(getter)(stock_code)
        if data:
            combined_data.update(data)
            
    return combined_data

if __name__ == "__main__":
    setup_logging_and_cleanup()
    process_stock_list("./template/stock_list.xlsx")
    logging.info("종료")